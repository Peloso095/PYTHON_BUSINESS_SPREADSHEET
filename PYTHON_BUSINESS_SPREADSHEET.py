# ------------------------------------------------------------------------------
# Copyright (c) 2025, [Rafael Peloso]
# Todos os direitos reservados.
#
# Licença Comercial - Somente compradores do produto podem modificar ou redistribuir
# o código e a planilha. A modificação ou distribuição do código sem a compra do 
# produto ou sem autorização expressa do autor é estritamente proibida.
#
# Este código é fornecido "NO ESTADO EM QUE SE ENCONTRA", sem garantia de qualquer
# tipo, expressa ou implícita, incluindo, mas não se limitando às garantias de
# comercialização, adequação a um propósito específico e não violação.
# ------------------------------------------------------------------------------

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# Criar uma nova planilha
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Vendas"

# Adicionar cabeçalhos
headers = [
    "Vendedor", "Vendas", "Comissão (%)", "Comissão (R$)", "Salário Total (R$)", "Desconto Vendas (%)", "Desconto Vendas (R$)", 
    "Imposto (%)", "Imposto (R$)", "Bônus (R$)", "Total Vendas Acumulado (R$)", "Meta de Vendas Atingida?", "13º Salário (R$)",
    "Horas Extras (R$)", "Desconto INSS (R$)", "Desconto IR (R$)", "Valor a Pagar após Descontos (R$)", "Valor Médio de Vendas (R$)",
    "Maior Venda (R$)", "Menor Venda (R$)", "Número de Vendas", "Comissão Progressiva (R$)", "Comissão Fixa (R$)", 
    "Saldo Comissão (R$)", "Meta de Vendas (%)", "Saldo Final após Venda (R$)", "Comissão por Produto (R$)", 
    "Vendas com Maior Comissão (R$)", "Vendas com Menor Comissão (R$)", "Vendas por Categoria", "Custo por Produto (R$)",
    "Lucro por Venda (R$)", "Margem de Lucro (%)", "Comissão por Tipo de Cliente (R$)", "Rendimento do Vendedor (R$)",
    "Lucro Líquido após Vendas (R$)", "Comissões por Promoção (R$)", "Comissão Extra por Novo Produto (R$)",
    "Comissão Extra por Novo Cliente (R$)", "Vendas por Período", "Desempenho comparado ao Mês Anterior (%)", 
    "Vendas por Região", "Comissão por Equipe (R$)", "Desconto Cliente Fiel (R$)"
]

# Inserir os cabeçalhos na planilha com formatação avançada
for col_num, header in enumerate(headers, start=1):
    col_letter = get_column_letter(col_num)
    cell = sheet[f"{col_letter}1"]
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF", size=12)  # Fonte em negrito e branca
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Cor de fundo azul
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(bottom=Side(style='thin', color="000000"))  # Adicionando borda fina inferior

# Ajustar o tamanho das colunas
for col_num in range(1, len(headers) + 1):
    col_letter = get_column_letter(col_num)
    max_length = 0
    for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[col_letter].width = adjusted_width

# Adicionar uma imagem de logo (exemplo)
# img = Image("logo.png")  # Exemplo: se você tiver uma imagem no seu diretório
# sheet.add_image(img, 'A1')

# Dados fictícios
dados = [
    ["Ana", 1000, 5, 10, 2000],
    ["João", 1500, 7, 10, 2500],
    ["Carlos", 2000, 8, 12, 3000],
]

# Preencher os dados na planilha com formatação condicional
for i, (vendedor, vendas, comissao, desconto, salario) in enumerate(dados, start=2):
    sheet[f"A{i}"] = vendedor
    sheet[f"B{i}"] = vendas
    sheet[f"C{i}"] = comissao
    sheet[f"D{i}"] = f"=B{i}*C{i}/100"
    sheet[f"E{i}"] = f"=B{i}+D{i}"
    sheet[f"F{i}"] = desconto
    sheet[f"G{i}"] = f"=B{i}*F{i}/100"
    sheet[f"H{i}"] = 5  # Exemplo de imposto 5%
    sheet[f"I{i}"] = f"=B{i}*H{i}/100"
    sheet[f"J{i}"] = 100  # Bônus fixo de 100
    sheet[f"K{i}"] = f"=B{i}"  # Total de vendas acumulado simples
    sheet[f"L{i}"] = f"=SE(B{i}>2000;\"Sim\";\"Não\")"  # Meta de vendas
    sheet[f"M{i}"] = f"=E{i}/12"  # 13º salário proporcional
    sheet[f"N{i}"] = 50  # Exemplo de horas extras 50
    sheet[f"O{i}"] = f"=E{i}*0.11"  # Desconto INSS 11%
    sheet[f"P{i}"] = f"=E{i}*0.15"  # Desconto IR 15%
    sheet[f"Q{i}"] = f"=E{i}-O{i}-P{i}"  # Valor após descontos
    sheet[f"R{i}"] = f"=B{i}/10"  # Valor médio de vendas
    sheet[f"S{i}"] = f"=MÁXIMO(B2:B{len(dados)+1})"  # Maior venda
    sheet[f"T{i}"] = f"=MÍNIMO(B2:B{len(dados)+1})"  # Menor venda
    sheet[f"U{i}"] = f"=CONT.VALORES(B2:B{len(dados)+1})"  # Número de vendas
    sheet[f"V{i}"] = f"=SE(B{i}>1000;B{i}*0.1;0)"  # Comissão progressiva
    sheet[f"W{i}"] = 50  # Comissão fixa de 50
    sheet[f"X{i}"] = f"=D{i}+W{i}"  # Saldo de comissão
    sheet[f"Y{i}"] = f"=SE(B{i}>2000;(B{i}/2000)*100;0)"  # Meta de vendas em %
    sheet[f"Z{i}"] = f"=B{i}+V{i}"  # Saldo final após venda
    sheet[f"AA{i}"] = f"=B{i}*0.05"  # Comissão por produto
    sheet[f"AB{i}"] = f"=SE(D{i}=MÁXIMO(D2:D{len(dados)+1});B{i};0)"  # Vendas com maior comissão
    sheet[f"AC{i}"] = f"=SE(D{i}=MÍNIMO(D2:D{len(dados)+1});B{i};0)"  # Vendas com menor comissão
    sheet[f"AD{i}"] = "Produto A"  # Categoria de produto (exemplo)
    sheet[f"AE{i}"] = f"=B{i}/10"  # Custo por produto
    sheet[f"AF{i}"] = f"=B{i}-AE{i}"  # Lucro por venda
    sheet[f"AG{i}"] = f"=AF{i}/B{i}"  # Margem de lucro
    sheet[f"AH{i}"] = f"=B{i}*0.05"  # Comissão por tipo de cliente
    sheet[f"AI{i}"] = f"=E{i}"  # Rendimento do vendedor
    sheet[f"AJ{i}"] = f"=B{i}-C{i}"  # Lucro líquido após vendas
    sheet[f"AK{i}"] = f"=B{i}*0.02"  # Comissões por promoção
    sheet[f"AL{i}"] = f"=B{i}*0.03"  # Comissão extra por novo produto
    sheet[f"AM{i}"] = f"=B{i}*0.04"  # Comissão extra por novo cliente
    sheet[f"AN{i}"] = "Mensal"  # Período de vendas
    sheet[f"AO{i}"] = f"=SE(B{i}>B{i-1};\"Melhor\";\"Pior\")"  # Desempenho comparado ao mês anterior
    sheet[f"AP{i}"] = "Região X"  # Vendas por região
    sheet[f"AQ{i}"] = f"=D{i}+F{i}"  # Comissão por equipe
    sheet[f"AR{i}"] = f"=SE(B{i}>1500;B{i}*0.05;0)"  # Desconto cliente fiel

# Congelar a primeira linha
sheet.freeze_panes = "A2"

# Salvar a planilha
wb.save("PYTHON_BUSINESS_SPREADSHEET.xlsx")
